#!/usr/bin/env python3
"""
Report Generator
Generates various report formats from reconciliation data
"""

import json
from datetime import datetime
from decimal import Decimal
from typing import Dict, List
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reconciliation_engine import ReconciliationReport, ReconciliationEngine


class DecimalEncoder(json.JSONEncoder):
    """JSON encoder that handles Decimal types"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


class ReportGenerator:
    """Generates reports in various formats"""

    def __init__(self, engine: ReconciliationEngine):
        self.engine = engine

    def generate_summary_text(self, report: ReconciliationReport) -> str:
        """Generate a text summary report"""
        lines = []
        lines.append("=" * 80)
        lines.append("RECONCILIATION SUMMARY REPORT")
        lines.append("=" * 80)
        lines.append(f"Generated: {report.report_date.strftime('%Y-%m-%d %H:%M:%S')}")

        if report.date_range_start and report.date_range_end:
            lines.append(f"Date Range: {report.date_range_start.strftime('%Y-%m-%d')} to {report.date_range_end.strftime('%Y-%m-%d')}")

        lines.append("")
        lines.append("FINANCIAL SUMMARY")
        lines.append("-" * 80)
        lines.append(f"Total Remittances Received:     ${report.total_remittances_received:>15,.2f}")
        lines.append(f"Total Paid to Drivers:          ${report.total_paid_to_drivers:>15,.2f}")
        lines.append(f"Total Scheduled Amount:         ${report.total_scheduled_amount:>15,.2f}")
        lines.append(f"Difference (Received - Paid):   ${(report.total_remittances_received - report.total_paid_to_drivers):>15,.2f}")

        lines.append("")
        lines.append("RECONCILIATION STATISTICS")
        lines.append("-" * 80)
        lines.append(f"Fully Matched Transactions:     {len(report.full_matches):>10}")
        lines.append(f"Partial Matches:                {len(report.partial_matches):>10}")
        lines.append(f"Missing Bank Transactions:      {len(report.missing_bank_transactions):>10}")
        lines.append(f"Orphan Bank Transactions:       {len(report.orphan_bank_transactions):>10}")
        lines.append(f"Orphan Payment Remittances:     {len(report.orphan_payments):>10}")
        lines.append(f"Amount Discrepancies:           {len(report.amount_discrepancies):>10}")

        # Driver summaries
        if report.driver_summaries:
            lines.append("")
            lines.append("DRIVER SUMMARIES")
            lines.append("-" * 80)
            lines.append(f"{'Driver':<15} {'Scheduled':<12} {'Paid':<12} {'Unpaid':<12} {'Difference':<12}")
            lines.append("-" * 80)

            for driver, summary in sorted(report.driver_summaries.items()):
                scheduled = summary['scheduled_amount']
                paid = summary['paid_amount']
                unpaid = summary['unpaid_amount']
                diff = scheduled - paid

                lines.append(
                    f"{driver:<15} "
                    f"${scheduled:>10,.2f} "
                    f"${paid:>10,.2f} "
                    f"${unpaid:>10,.2f} "
                    f"${diff:>10,.2f}"
                )

        # Missing bank transactions (unpaid loads)
        if report.missing_bank_transactions:
            lines.append("")
            lines.append("UNPAID LOADS (Missing Bank Transactions)")
            lines.append("-" * 80)
            lines.append(f"{'Date':<12} {'Driver':<15} {'Company':<20} {'Load #':<15} {'Amount':<12}")
            lines.append("-" * 80)

            for entry in sorted(report.missing_bank_transactions, key=lambda x: x.date):
                lines.append(
                    f"{entry.date.strftime('%Y-%m-%d'):<12} "
                    f"{entry.driver:<15} "
                    f"{entry.company:<20} "
                    f"{entry.load_number:<15} "
                    f"${entry.amount:>10,.2f}" if entry.amount else "N/A"
                )

        # Orphan bank transactions
        if report.orphan_bank_transactions:
            lines.append("")
            lines.append("ORPHAN BANK TRANSACTIONS (Paid but not scheduled)")
            lines.append("-" * 80)
            lines.append(f"{'Date':<12} {'Recipient':<20} {'Description':<30} {'Amount':<12}")
            lines.append("-" * 80)

            for trans in sorted(report.orphan_bank_transactions, key=lambda x: x.transaction_date):
                lines.append(
                    f"{trans.transaction_date.strftime('%Y-%m-%d'):<12} "
                    f"{trans.recipient:<20} "
                    f"{trans.description:<30} "
                    f"${trans.amount:>10,.2f}"
                )

        lines.append("")
        lines.append("=" * 80)

        return "\n".join(lines)

    def generate_detailed_text(self, report: ReconciliationReport) -> str:
        """Generate a detailed text report with all matches"""
        lines = []
        lines.append("=" * 80)
        lines.append("DETAILED RECONCILIATION REPORT")
        lines.append("=" * 80)

        # Add summary first
        lines.append(self.generate_summary_text(report))

        # Full matches
        if report.full_matches:
            lines.append("")
            lines.append("FULLY MATCHED TRANSACTIONS")
            lines.append("-" * 80)

            for match in sorted(report.full_matches,
                              key=lambda x: x.driver_entry.date if x.driver_entry else datetime.min):
                if match.driver_entry:
                    lines.append(f"\nSchedule Date: {match.driver_entry.date.strftime('%Y-%m-%d')}")
                    lines.append(f"  Driver: {match.driver_entry.driver}")
                    lines.append(f"  Company: {match.driver_entry.company}")
                    lines.append(f"  Load #: {match.driver_entry.load_number}")
                    lines.append(f"  Scheduled Amount: ${match.driver_entry.amount:,.2f}" if match.driver_entry.amount else "  Scheduled Amount: N/A")

                    # For Ready direct payments, show the Ready payment date
                    if match.match_type == "READY_DIRECT_PAYMENT":
                        if match.driver_entry.date_paid:
                            lines.append(f"  Ready Payment Date: {match.driver_entry.date_paid.strftime('%Y-%m-%d')}")
                            lines.append(f"  Days to Payment: {(match.driver_entry.date_paid - match.driver_entry.date).days}")
                        lines.append(f"  Note: Ready direct payment (no driver payment needed)")
                    elif match.bank_transaction:
                        lines.append(f"  Payment Date: {match.bank_transaction.transaction_date.strftime('%Y-%m-%d')}")
                        lines.append(f"  Paid Amount: ${match.bank_transaction.amount:,.2f}")
                        lines.append(f"  Days to Payment: {(match.bank_transaction.transaction_date - match.driver_entry.date).days}")

        lines.append("")
        lines.append("=" * 80)

        return "\n".join(lines)

    def generate_json_report(self, report: ReconciliationReport) -> str:
        """Generate a JSON report for API/web interface"""
        data = {
            'report_date': report.report_date,
            'date_range': {
                'start': report.date_range_start,
                'end': report.date_range_end,
            },
            'financial_summary': {
                'total_remittances_received': report.total_remittances_received,
                'total_paid_to_drivers': report.total_paid_to_drivers,
                'total_scheduled_amount': report.total_scheduled_amount,
                'difference': report.total_remittances_received - report.total_paid_to_drivers,
            },
            'statistics': {
                'full_matches': len(report.full_matches),
                'partial_matches': len(report.partial_matches),
                'missing_bank_transactions': len(report.missing_bank_transactions),
                'orphan_bank_transactions': len(report.orphan_bank_transactions),
                'orphan_payments': len(report.orphan_payments),
                'amount_discrepancies': len(report.amount_discrepancies),
            },
            'driver_summaries': report.driver_summaries,
            'paid_loads': [
                {
                    'date': match.driver_entry.date if match.driver_entry else None,
                    'driver': match.driver_entry.driver if match.driver_entry else None,
                    'company': match.driver_entry.company if match.driver_entry else None,
                    'load_number': match.driver_entry.load_number if match.driver_entry else None,
                    'amount': match.driver_entry.amount if match.driver_entry else None,
                    'pickup': match.driver_entry.pickup if match.driver_entry else None,
                    'dropoff': match.driver_entry.dropoff if match.driver_entry else None,
                    'date_paid': match.driver_entry.date_paid if (match.driver_entry and match.driver_entry.date_paid) else (match.bank_transaction.transaction_date if match.bank_transaction else None),
                    'payment_type': match.match_type,
                }
                for match in report.full_matches
            ],
            'unpaid_loads': [
                {
                    'date': entry.date,
                    'driver': entry.driver,
                    'company': entry.company,
                    'load_number': entry.load_number,
                    'amount': entry.amount,
                    'pickup': entry.pickup,
                    'dropoff': entry.dropoff,
                }
                for entry in report.missing_bank_transactions
            ],
            'orphan_bank_transactions': [
                {
                    'date': trans.transaction_date,
                    'recipient': trans.recipient,
                    'amount': trans.amount,
                    'description': trans.description,
                }
                for trans in report.orphan_bank_transactions
            ],
        }

        return json.dumps(data, indent=2, cls=DecimalEncoder)

    def generate_csv_driver_summary(self, report: ReconciliationReport) -> str:
        """Generate CSV format driver summary"""
        lines = []
        lines.append("Driver,Scheduled Loads,Scheduled Amount,Paid Loads,Paid Amount,Unpaid Loads,Unpaid Amount,Difference")

        for driver, summary in sorted(report.driver_summaries.items()):
            lines.append(
                f"{driver},"
                f"{summary['scheduled_loads']},"
                f"{summary['scheduled_amount']},"
                f"{summary['paid_loads']},"
                f"{summary['paid_amount']},"
                f"{summary['unpaid_loads']},"
                f"{summary['unpaid_amount']},"
                f"{summary['scheduled_amount'] - summary['paid_amount']}"
            )

        return "\n".join(lines)

    def generate_html_report(self, report: ReconciliationReport) -> str:
        """Generate HTML report for web display"""
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Reconciliation Report - {report.report_date.strftime('%Y-%m-%d')}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }}
        h1, h2 {{
            color: #333;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            padding: 10px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #4CAF50;
            color: white;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .summary-box {{
            background-color: #e7f3fe;
            border-left: 4px solid #2196F3;
            padding: 15px;
            margin: 20px 0;
        }}
        .warning {{
            background-color: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 15px;
            margin: 20px 0;
        }}
        .amount {{
            text-align: right;
            font-family: monospace;
        }}
        .positive {{
            color: green;
        }}
        .negative {{
            color: red;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Reconciliation Report</h1>
        <p><strong>Generated:</strong> {report.report_date.strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p><strong>Date Range:</strong> {report.date_range_start.strftime('%Y-%m-%d') if report.date_range_start else 'N/A'} to {report.date_range_end.strftime('%Y-%m-%d') if report.date_range_end else 'N/A'}</p>

        <div class="summary-box">
            <h2>Financial Summary</h2>
            <table>
                <tr>
                    <td>Total Remittances Received:</td>
                    <td class="amount">${report.total_remittances_received:,.2f}</td>
                </tr>
                <tr>
                    <td>Total Paid to Drivers:</td>
                    <td class="amount">${report.total_paid_to_drivers:,.2f}</td>
                </tr>
                <tr>
                    <td>Total Scheduled Amount:</td>
                    <td class="amount">${report.total_scheduled_amount:,.2f}</td>
                </tr>
                <tr>
                    <td><strong>Difference (Received - Paid):</strong></td>
                    <td class="amount {'positive' if (report.total_remittances_received - report.total_paid_to_drivers) >= 0 else 'negative'}">
                        <strong>${(report.total_remittances_received - report.total_paid_to_drivers):,.2f}</strong>
                    </td>
                </tr>
            </table>
        </div>

        <h2>Driver Summaries</h2>
        <table>
            <thead>
                <tr>
                    <th>Driver</th>
                    <th>Scheduled</th>
                    <th>Paid</th>
                    <th>Unpaid</th>
                    <th>Difference</th>
                </tr>
            </thead>
            <tbody>
        """

        for driver, summary in sorted(report.driver_summaries.items()):
            diff = summary['scheduled_amount'] - summary['paid_amount']
            diff_class = 'negative' if diff < 0 else 'positive' if diff > 0 else ''

            html += f"""
                <tr>
                    <td>{driver}</td>
                    <td class="amount">${summary['scheduled_amount']:,.2f}</td>
                    <td class="amount">${summary['paid_amount']:,.2f}</td>
                    <td class="amount">${summary['unpaid_amount']:,.2f}</td>
                    <td class="amount {diff_class}">${diff:,.2f}</td>
                </tr>
            """

        html += """
            </tbody>
        </table>
        """

        # Unpaid loads
        if report.missing_bank_transactions:
            html += """
        <div class="warning">
            <h2>Unpaid Loads</h2>
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Driver</th>
                        <th>Company</th>
                        <th>Load #</th>
                        <th>Amount</th>
                    </tr>
                </thead>
                <tbody>
            """

            for entry in sorted(report.missing_bank_transactions, key=lambda x: x.date):
                amount_str = f"${entry.amount:,.2f}" if entry.amount else "N/A"
                html += f"""
                    <tr>
                        <td>{entry.date.strftime('%Y-%m-%d')}</td>
                        <td>{entry.driver}</td>
                        <td>{entry.company}</td>
                        <td>{entry.load_number}</td>
                        <td class="amount">{amount_str}</td>
                    </tr>
                """

            html += """
                </tbody>
            </table>
        </div>
            """

        html += """
    </div>
</body>
</html>
        """

        return html

    def generate_comprehensive_excel_report(self, report: ReconciliationReport) -> Workbook:
        """
        Generate comprehensive load tracking Excel report

        One row per load with all payment information:
        - Load Company, Load#, Load Amount/Date
        - Driver, Driver Pay Amount/Date
        - Factoring Agent/Payment Amount/Date
        - Deposit Amount/Date
        - Matched Using/Date
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Load Tracking"

        # Define headers
        headers = [
            'Load Date',
            'Company',
            'Load Number',
            'Pickup',
            'Dropoff',
            'Driver',
            'Load Amount',
            'Driver Pay Amount',
            'Driver Pay Date',
            'Days to Driver Pay',
            'Factoring Agent',
            'Factoring Amount',
            'Factoring Date',
            'Deposit Amount',
            'Deposit Date',
            'Match Status',
            'Matched Using',
            'Notes'
        ]

        # Write headers with styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Set column widths
        column_widths = {
            'A': 12,  # Load Date
            'B': 20,  # Company
            'C': 15,  # Load Number
            'D': 12,  # Pickup
            'E': 12,  # Dropoff
            'F': 15,  # Driver
            'G': 13,  # Load Amount
            'H': 15,  # Driver Pay Amount
            'I': 15,  # Driver Pay Date
            'J': 12,  # Days to Driver Pay
            'K': 18,  # Factoring Agent
            'L': 15,  # Factoring Amount
            'M': 15,  # Factoring Date
            'N': 15,  # Deposit Amount
            'O': 15,  # Deposit Date
            'P': 20,  # Match Status
            'Q': 25,  # Matched Using
            'R': 30,  # Notes
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Build comprehensive load list
        # We'll iterate through all driver schedule entries (loads)
        # and match them with payment data

        row_num = 2

        # Create lookup dictionaries for efficient matching
        # Map load_number -> PaymentRemittance invoice
        factoring_by_load = {}
        for remittance in self.engine.data.payment_remittances:
            for invoice in remittance.invoices:
                load_num = invoice.get('vin', '')  # Using VIN as identifier
                if load_num:
                    factoring_by_load[load_num] = {
                        'agent': 'Cox Automotive',
                        'amount': invoice['amount_paid'],
                        'date': remittance.payment_date,
                    }

        # Map load_number -> ReadyStatement
        ready_by_load = {}
        for statement in self.engine.data.ready_statements:
            ready_by_load[statement.invoice_number] = statement

        # Process all loads from driver schedules
        all_loads = sorted(self.engine.data.driver_schedules, key=lambda x: x.date)

        for load in all_loads:
            # Find if this load has been matched in the reconciliation
            match_info = None
            match_status = "Unmatched"
            matched_using = ""
            notes = ""

            # Find the match in the report
            for match in report.full_matches:
                if match.driver_entry and id(match.driver_entry) == id(load):
                    match_info = match
                    if match.match_type == "READY_DIRECT_PAYMENT":
                        match_status = "Ready Direct Payment"
                        matched_using = "Ready Statement"
                    elif match.match_type == "FULL":
                        match_status = "Fully Matched"
                        matched_using = "Driver + Amount + Date"
                    break

            # Check if in missing bank transactions
            if not match_info:
                for missing in report.missing_bank_transactions:
                    if id(missing) == id(load):
                        match_status = "Unpaid"
                        matched_using = "No bank transaction found"
                        notes = "Driver payment pending"
                        break

            # Basic load information
            ws.cell(row=row_num, column=1, value=load.date.strftime('%Y-%m-%d') if load.date else '')
            ws.cell(row=row_num, column=2, value=load.company)
            ws.cell(row=row_num, column=3, value=load.load_number)
            ws.cell(row=row_num, column=4, value=load.pickup)
            ws.cell(row=row_num, column=5, value=load.dropoff)
            ws.cell(row=row_num, column=6, value=load.driver)
            ws.cell(row=row_num, column=7, value=float(load.amount) if load.amount else '')

            # Driver payment information
            if match_info and match_info.bank_transaction:
                ws.cell(row=row_num, column=8, value=float(match_info.bank_transaction.amount))
                ws.cell(row=row_num, column=9, value=match_info.bank_transaction.transaction_date.strftime('%Y-%m-%d'))
                days_diff = (match_info.bank_transaction.transaction_date - load.date).days
                ws.cell(row=row_num, column=10, value=days_diff)
            elif load.date_paid:
                # For Ready loads with date_paid set
                ws.cell(row=row_num, column=8, value='N/A - Direct Payment')
                ws.cell(row=row_num, column=9, value=load.date_paid.strftime('%Y-%m-%d'))
                days_diff = (load.date_paid - load.date).days
                ws.cell(row=row_num, column=10, value=days_diff)
            else:
                ws.cell(row=row_num, column=8, value='')
                ws.cell(row=row_num, column=9, value='')
                ws.cell(row=row_num, column=10, value='')

            # Factoring information (Cox or Ready)
            if load.company.upper() == "READY" and load.load_number in ready_by_load:
                ready_stmt = ready_by_load[load.load_number]
                ws.cell(row=row_num, column=11, value='Ready Logistics')
                ws.cell(row=row_num, column=12, value=float(ready_stmt.payment_amount))
                ws.cell(row=row_num, column=13, value=ready_stmt.payment_date.strftime('%Y-%m-%d'))
            else:
                # Try to find in Cox factoring by VIN or load number
                # Note: This mapping is complex and may need business logic
                ws.cell(row=row_num, column=11, value='Cox Automotive')
                ws.cell(row=row_num, column=12, value='')  # Would need VIN mapping
                ws.cell(row=row_num, column=13, value='')

            # Deposit information (from Ready statements for Ready loads)
            if load.company.upper() == "READY" and load.load_number in ready_by_load:
                ready_stmt = ready_by_load[load.load_number]
                ws.cell(row=row_num, column=14, value=float(ready_stmt.payment_amount))
                ws.cell(row=row_num, column=15, value=ready_stmt.payment_date.strftime('%Y-%m-%d'))
            else:
                # Deposit data for Cox loads would come from bank deposits
                # This requires additional mapping logic
                ws.cell(row=row_num, column=14, value='')
                ws.cell(row=row_num, column=15, value='')

            # Match status and method
            ws.cell(row=row_num, column=16, value=match_status)
            ws.cell(row=row_num, column=17, value=matched_using)
            ws.cell(row=row_num, column=18, value=notes)

            # Apply styling to data rows
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Format currency columns
                if col in [7, 8, 12, 14]:  # Amount columns
                    if cell.value and cell.value != '':
                        cell.number_format = '$#,##0.00'

                # Color code by match status
                if match_status == "Fully Matched":
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif match_status == "Ready Direct Payment":
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                elif match_status == "Unpaid":
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

            row_num += 1

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

        return wb

    def save_reports(self, report: ReconciliationReport, output_dir: str = "reports"):
        """Save all report formats to files"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        timestamp = report.report_date.strftime('%Y%m%d_%H%M%S')

        # Save text summary
        summary_file = output_path / f"summary_{timestamp}.txt"
        with open(summary_file, 'w') as f:
            f.write(self.generate_summary_text(report))

        # Save detailed text
        detailed_file = output_path / f"detailed_{timestamp}.txt"
        with open(detailed_file, 'w') as f:
            f.write(self.generate_detailed_text(report))

        # Save JSON
        json_file = output_path / f"report_{timestamp}.json"
        with open(json_file, 'w') as f:
            f.write(self.generate_json_report(report))

        # Save CSV
        csv_file = output_path / f"driver_summary_{timestamp}.csv"
        with open(csv_file, 'w') as f:
            f.write(self.generate_csv_driver_summary(report))

        # Save HTML
        html_file = output_path / f"report_{timestamp}.html"
        with open(html_file, 'w') as f:
            f.write(self.generate_html_report(report))

        # Save Comprehensive Excel Load Tracking Report
        excel_file = output_path / f"load_tracking_{timestamp}.xlsx"
        wb = self.generate_comprehensive_excel_report(report)
        wb.save(str(excel_file))

        return {
            'summary': str(summary_file),
            'detailed': str(detailed_file),
            'json': str(json_file),
            'csv': str(csv_file),
            'html': str(html_file),
            'excel': str(excel_file),
        }
